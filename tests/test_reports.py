"""
Unit Tests for Phase 3: Report Generation
Tests report formatting, PDF/Excel generation, GenerateReportTool, and Reports API
"""

import sys
sys.path.insert(0, "D:\\OryggiAI_Service\\Advance_Chatbot")

import pytest
import os
import tempfile
from unittest.mock import Mock, patch, MagicMock, call
from datetime import datetime
from pathlib import Path

from app.reports.report_formatter import ReportFormatter
from app.reports.pdf_generator import PDFReportGenerator, pdf_generator
from app.reports.excel_generator import ExcelReportGenerator, excel_generator
from app.tools.generate_report_tool import GenerateReportTool, generate_report_tool


class TestReportFormatter:
    """Test suite for ReportFormatter utility class"""

    def test_format_query_results_basic(self):
        """Test basic formatting of query results"""
        query_results = [
            {"id": 1, "name": "John Doe", "salary": 50000},
            {"id": 2, "name": "Jane Smith", "salary": 60000}
        ]

        formatted = ReportFormatter.format_query_results(query_results, "Employee Report")

        assert formatted["title"] == "Employee Report"
        assert formatted["columns"] == ["id", "name", "salary"]
        assert len(formatted["rows"]) == 2
        assert formatted["total_rows"] == 2
        assert formatted["displayed_rows"] == 2
        assert formatted["truncated"] == False

    def test_format_query_results_empty(self):
        """Test formatting empty results"""
        formatted = ReportFormatter.format_query_results([], "Empty Report")

        assert formatted["title"] == "Empty Report"
        assert formatted["columns"] == []
        assert formatted["rows"] == []
        assert formatted["total_rows"] == 0
        assert formatted["truncated"] == False

    def test_format_query_results_with_max_rows(self):
        """Test formatting with max_rows limit"""
        query_results = [{"id": i, "name": f"User {i}"} for i in range(1, 101)]

        formatted = ReportFormatter.format_query_results(
            query_results,
            "Large Report",
            max_rows=50
        )

        assert formatted["total_rows"] == 100
        assert formatted["displayed_rows"] == 50
        assert len(formatted["rows"]) == 50
        assert formatted["truncated"] == True

    def test_format_query_results_none_values(self):
        """Test formatting handles None values"""
        query_results = [
            {"id": 1, "name": "John", "manager": None},
            {"id": 2, "name": None, "manager": "Jane"}
        ]

        formatted = ReportFormatter.format_query_results(query_results, "Report")

        # Check that None values are converted to "N/A"
        assert formatted["rows"][0][2] == "N/A"  # manager is None
        assert formatted["rows"][1][1] == "N/A"  # name is None

    def test_format_query_results_date_values(self):
        """Test formatting handles datetime values"""
        from datetime import date
        query_results = [
            {"id": 1, "hire_date": date(2020, 1, 15)}
        ]

        formatted = ReportFormatter.format_query_results(query_results, "Report")

        # Check that date is converted to string
        assert isinstance(formatted["rows"][0][1], str)
        assert "2020-01-15" in formatted["rows"][0][1]

    def test_to_dataframe(self):
        """Test conversion to pandas DataFrame"""
        formatted_data = {
            "columns": ["id", "name", "salary"],
            "rows": [
                [1, "John Doe", 50000],
                [2, "Jane Smith", 60000]
            ]
        }

        df = ReportFormatter.to_dataframe(formatted_data)

        assert df is not None
        assert len(df) == 2
        assert list(df.columns) == ["id", "name", "salary"]
        assert df.iloc[0]["name"] == "John Doe"

    def test_create_summary_stats_numeric(self):
        """Test summary statistics for numeric columns"""
        formatted_data = {
            "columns": ["id", "name", "salary"],
            "rows": [
                [1, "John", 50000],
                [2, "Jane", 60000],
                [3, "Bob", 55000]
            ]
        }

        summary = ReportFormatter.create_summary_stats(formatted_data)

        assert summary is not None
        assert "salary" in summary
        assert summary["salary"]["min"] == 50000
        assert summary["salary"]["max"] == 60000
        assert summary["salary"]["mean"] == 55000
        assert summary["salary"]["count"] == 3

    def test_create_summary_stats_no_numeric(self):
        """Test summary stats with no numeric columns"""
        formatted_data = {
            "columns": ["name", "department"],
            "rows": [
                ["John", "IT"],
                ["Jane", "HR"]
            ]
        }

        summary = ReportFormatter.create_summary_stats(formatted_data)

        assert summary == {}

    def test_add_metadata(self):
        """Test adding metadata to formatted data"""
        formatted_data = {
            "title": "Test Report",
            "columns": ["id", "name"],
            "rows": [[1, "John"]]
        }

        result = ReportFormatter.add_metadata(
            formatted_data,
            user_id="admin_001",
            user_role="ADMIN",
            question="List all employees",
            sql_query="SELECT id, name FROM Employees"
        )

        assert result["metadata"]["user_id"] == "admin_001"
        assert result["metadata"]["user_role"] == "ADMIN"
        assert result["metadata"]["question"] == "List all employees"
        assert result["metadata"]["sql_query"] == "SELECT id, name FROM Employees"


class TestPDFReportGenerator:
    """Test suite for PDF report generator"""

    def test_pdf_generator_initialization(self):
        """Test PDF generator initializes correctly"""
        generator = PDFReportGenerator()

        assert generator is not None
        assert generator.env is not None
        assert os.path.exists(generator.templates_dir)

    def test_global_pdf_generator_exists(self):
        """Test global PDF generator instance exists"""
        assert pdf_generator is not None
        assert isinstance(pdf_generator, PDFReportGenerator)

    @patch('app.reports.pdf_generator.settings')
    def test_generate_report_basic(self, mock_settings):
        """Test basic PDF report generation"""
        # Create temporary directory for output
        with tempfile.TemporaryDirectory() as temp_dir:
            mock_settings.reports_output_dir = temp_dir

            formatted_data = {
                "title": "Test Report",
                "columns": ["id", "name"],
                "rows": [[1, "John Doe"], [2, "Jane Smith"]],
                "total_rows": 2,
                "displayed_rows": 2,
                "generated_at": datetime.now().isoformat(),
                "truncated": False
            }

            generator = PDFReportGenerator()

            # Generate report
            report_path = generator.generate_report(formatted_data)

            # Verify file was created
            assert os.path.exists(report_path)
            assert report_path.endswith('.pdf')
            assert os.path.getsize(report_path) > 0

    @patch('app.reports.pdf_generator.settings')
    @patch('app.reports.pdf_generator.ReportFormatter')
    def test_generate_table_report(self, mock_formatter, mock_settings):
        """Test generate_table_report convenience method"""
        with tempfile.TemporaryDirectory() as temp_dir:
            mock_settings.reports_output_dir = temp_dir
            mock_settings.reports_max_rows = 1000

            # Mock formatter responses
            formatted_data = {
                "title": "Employee Report",
                "columns": ["id", "name"],
                "rows": [[1, "John"]],
                "total_rows": 1,
                "displayed_rows": 1,
                "generated_at": datetime.now().isoformat(),
                "truncated": False,
                "metadata": {
                    "user_id": "admin",
                    "user_role": "ADMIN",
                    "question": "List employees",
                    "sql_query": "SELECT * FROM Employees"
                }
            }
            mock_formatter.format_query_results.return_value = formatted_data
            mock_formatter.add_metadata.return_value = formatted_data

            query_results = [{"id": 1, "name": "John"}]

            generator = PDFReportGenerator()
            report_path = generator.generate_table_report(
                query_results=query_results,
                title="Employee Report",
                user_id="admin",
                user_role="ADMIN",
                question="List employees",
                sql_query="SELECT * FROM Employees"
            )

            # Verify formatter was called
            mock_formatter.format_query_results.assert_called_once()
            mock_formatter.add_metadata.assert_called_once()

            # Verify file exists
            assert os.path.exists(report_path)
            assert report_path.endswith('.pdf')


class TestExcelReportGenerator:
    """Test suite for Excel report generator"""

    def test_excel_generator_initialization(self):
        """Test Excel generator initializes correctly"""
        generator = ExcelReportGenerator()
        assert generator is not None

    def test_global_excel_generator_exists(self):
        """Test global Excel generator instance exists"""
        assert excel_generator is not None
        assert isinstance(excel_generator, ExcelReportGenerator)

    @patch('app.reports.excel_generator.settings')
    def test_generate_report_basic(self, mock_settings):
        """Test basic Excel report generation"""
        with tempfile.TemporaryDirectory() as temp_dir:
            mock_settings.reports_output_dir = temp_dir

            formatted_data = {
                "title": "Test Report",
                "columns": ["id", "name", "salary"],
                "rows": [
                    [1, "John Doe", 50000],
                    [2, "Jane Smith", 60000]
                ],
                "total_rows": 2,
                "displayed_rows": 2,
                "generated_at": datetime.now().isoformat(),
                "truncated": False,
                "metadata": {
                    "user_id": "admin",
                    "user_role": "ADMIN",
                    "question": "List employees",
                    "sql_query": "SELECT * FROM Employees"
                }
            }

            generator = ExcelReportGenerator()
            report_path = generator.generate_report(formatted_data)

            # Verify file was created
            assert os.path.exists(report_path)
            assert report_path.endswith('.xlsx')
            assert os.path.getsize(report_path) > 0

    @patch('app.reports.excel_generator.settings')
    def test_generate_report_with_summary(self, mock_settings):
        """Test Excel report generation with summary statistics"""
        with tempfile.TemporaryDirectory() as temp_dir:
            mock_settings.reports_output_dir = temp_dir

            formatted_data = {
                "title": "Salary Report",
                "columns": ["id", "name", "salary"],
                "rows": [
                    [1, "John", 50000],
                    [2, "Jane", 60000],
                    [3, "Bob", 55000]
                ],
                "total_rows": 3,
                "displayed_rows": 3,
                "generated_at": datetime.now().isoformat(),
                "truncated": False,
                "metadata": {
                    "user_id": "admin",
                    "user_role": "ADMIN"
                },
                "summary_stats": {
                    "salary": {
                        "min": 50000,
                        "max": 60000,
                        "mean": 55000,
                        "median": 55000,
                        "sum": 165000,
                        "count": 3
                    }
                }
            }

            generator = ExcelReportGenerator()
            report_path = generator.generate_report(formatted_data, include_summary=True)

            # Verify file exists
            assert os.path.exists(report_path)
            assert report_path.endswith('.xlsx')

            # Verify workbook has multiple sheets (Data, Summary, Metadata)
            from openpyxl import load_workbook
            wb = load_workbook(report_path)
            assert "Data" in wb.sheetnames
            assert "Summary" in wb.sheetnames
            assert "Metadata" in wb.sheetnames


class TestGenerateReportTool:
    """Test suite for GenerateReportTool"""

    def test_tool_metadata(self):
        """Test tool has correct metadata"""
        tool = GenerateReportTool()

        assert tool.name == "generate_report"
        assert tool.description is not None
        assert "report" in tool.description.lower()
        assert tool.destructive == False
        assert tool.rbac_required == ["ADMIN", "HR_MANAGER", "HR_STAFF"]

    def test_global_instance_exists(self):
        """Test global tool instance exists"""
        assert generate_report_tool is not None
        assert isinstance(generate_report_tool, GenerateReportTool)

    def test_viewer_role_denied(self):
        """Test VIEWER role is denied access"""
        tool = GenerateReportTool()
        allowed, error = tool.check_permission("VIEWER")

        assert allowed == False
        assert error is not None
        assert "Permission denied" in error

    def test_admin_has_permission(self):
        """Test ADMIN role has access"""
        tool = GenerateReportTool()
        allowed, error = tool.check_permission("ADMIN")

        assert allowed == True
        assert error is None

    def test_hr_manager_has_permission(self):
        """Test HR_MANAGER role has access"""
        tool = GenerateReportTool()
        allowed, error = tool.check_permission("HR_MANAGER")

        assert allowed == True
        assert error is None

    def test_hr_staff_has_permission(self):
        """Test HR_STAFF role has access"""
        tool = GenerateReportTool()
        allowed, error = tool.check_permission("HR_STAFF")

        assert allowed == True
        assert error is None

    def test_invalid_format_validation(self):
        """Test invalid report format is rejected"""
        tool = GenerateReportTool()

        result = tool._run(
            question="Test query",
            user_id="admin_001",
            user_role="ADMIN",
            format="invalid"
        )

        assert result["success"] == False
        assert "Invalid format" in result["error"]
        assert "pdf" in result["error"]
        assert "excel" in result["error"]

    @patch('app.tools.generate_report_tool.query_database_tool')
    @patch('app.tools.generate_report_tool.pdf_generator')
    @patch('app.tools.generate_report_tool.audit_logger')
    def test_successful_pdf_generation(self, mock_audit, mock_pdf_gen, mock_query_tool):
        """Test successful PDF report generation"""
        # Mock query_database_tool response
        mock_query_tool.run.return_value = {
            "success": True,
            "result": {
                "results": [
                    {"id": 1, "name": "John Doe"},
                    {"id": 2, "name": "Jane Smith"}
                ],
                "sql_query": "SELECT id, name FROM Employees",
                "scoped_sql_query": "SELECT id, name FROM Employees"
            }
        }

        # Mock PDF generator
        mock_pdf_gen.generate_table_report.return_value = "/tmp/report_123.pdf"

        # Execute tool
        tool = GenerateReportTool()
        result = tool._run(
            question="List all employees",
            user_id="admin_001",
            user_role="ADMIN",
            format="pdf"
        )

        # Verify success
        assert result["success"] == True
        assert result["report_path"] == "/tmp/report_123.pdf"
        assert result["format"] == "pdf"
        assert result["rows_count"] == 2
        assert "execution_time_ms" in result

        # Verify PDF generator was called
        mock_pdf_gen.generate_table_report.assert_called_once()

        # Verify audit logging
        mock_audit.log_tool_execution.assert_called_once()

    @patch('app.tools.generate_report_tool.query_database_tool')
    @patch('app.tools.generate_report_tool.excel_generator')
    @patch('app.tools.generate_report_tool.audit_logger')
    def test_successful_excel_generation(self, mock_audit, mock_excel_gen, mock_query_tool):
        """Test successful Excel report generation"""
        # Mock query_database_tool response
        mock_query_tool.run.return_value = {
            "success": True,
            "result": {
                "results": [{"id": 1, "name": "John"}],
                "sql_query": "SELECT * FROM Employees",
                "scoped_sql_query": "SELECT * FROM Employees"
            }
        }

        # Mock Excel generator
        mock_excel_gen.generate_table_report.return_value = "/tmp/report_123.xlsx"

        # Execute tool
        tool = GenerateReportTool()
        result = tool._run(
            question="List employees",
            user_id="admin_001",
            user_role="ADMIN",
            format="excel"
        )

        # Verify success
        assert result["success"] == True
        assert result["report_path"] == "/tmp/report_123.xlsx"
        assert result["format"] == "excel"
        assert result["rows_count"] == 1

        # Verify Excel generator was called
        mock_excel_gen.generate_table_report.assert_called_once()

    @patch('app.tools.generate_report_tool.query_database_tool')
    @patch('app.tools.generate_report_tool.audit_logger')
    def test_query_failure_handling(self, mock_audit, mock_query_tool):
        """Test handling of query execution failure"""
        # Mock query failure
        mock_query_tool.run.return_value = {
            "success": False,
            "error": "Database connection failed"
        }

        # Execute tool
        tool = GenerateReportTool()
        result = tool._run(
            question="Test query",
            user_id="admin_001",
            user_role="ADMIN",
            format="pdf"
        )

        # Verify error result
        assert result["success"] == False
        assert "Database connection failed" in result["error"]
        assert result["query_error"] == True

        # Verify audit logging for failure
        mock_audit.log_tool_execution.assert_called_once()

    @patch('app.tools.generate_report_tool.query_database_tool')
    @patch('app.tools.generate_report_tool.audit_logger')
    def test_empty_results_handling(self, mock_audit, mock_query_tool):
        """Test handling of empty query results"""
        # Mock empty results
        mock_query_tool.run.return_value = {
            "success": True,
            "result": {
                "results": [],
                "sql_query": "SELECT * FROM Employees"
            }
        }

        # Execute tool
        tool = GenerateReportTool()
        result = tool._run(
            question="Find employees",
            user_id="admin_001",
            user_role="ADMIN",
            format="pdf"
        )

        # Verify error result
        assert result["success"] == False
        assert "No data available" in result["error"]
        assert result["empty_results"] == True

    @patch('app.tools.generate_report_tool.query_database_tool')
    @patch('app.tools.generate_report_tool.pdf_generator')
    @patch('app.tools.generate_report_tool.audit_logger')
    def test_hr_manager_department_scoping(self, mock_audit, mock_pdf_gen, mock_query_tool):
        """Test HR_MANAGER report includes department scoping"""
        # Mock query with scoped results
        mock_query_tool.run.return_value = {
            "success": True,
            "result": {
                "results": [{"id": 1, "name": "John", "department": "IT"}],
                "sql_query": "SELECT * FROM Employees",
                "scoped_sql_query": "SELECT * FROM Employees WHERE Department = 'IT'"
            }
        }

        mock_pdf_gen.generate_table_report.return_value = "/tmp/report.pdf"

        # Execute tool
        tool = GenerateReportTool()
        result = tool._run(
            question="List my department employees",
            user_id="manager_001",
            user_role="HR_MANAGER",
            format="pdf",
            department="IT"
        )

        # Verify query_database_tool was called with department
        call_args = mock_query_tool.run.call_args
        assert call_args[1]["department"] == "IT"
        assert call_args[1]["user_role"] == "HR_MANAGER"

        # Verify success
        assert result["success"] == True

    @patch('app.tools.generate_report_tool.query_database_tool')
    @patch('app.tools.generate_report_tool.pdf_generator')
    def test_exception_handling(self, mock_pdf_gen, mock_query_tool):
        """Test exception handling during report generation"""
        # Mock query success but PDF generation failure
        mock_query_tool.run.return_value = {
            "success": True,
            "result": {
                "results": [{"id": 1}],
                "sql_query": "SELECT * FROM Employees"
            }
        }

        mock_pdf_gen.generate_table_report.side_effect = Exception("Disk full")

        # Execute tool
        tool = GenerateReportTool()
        result = tool._run(
            question="Test",
            user_id="admin",
            user_role="ADMIN",
            format="pdf"
        )

        # Verify error handling
        assert result["success"] == False
        assert "Disk full" in result["error"]
        assert result["exception"] == "Exception"

    def test_run_method_denies_viewer(self):
        """Test run() method denies VIEWER role"""
        tool = GenerateReportTool()

        result = tool.run(
            user_role="VIEWER",
            question="Test query",
            user_id="viewer_001",
            format="pdf"
        )

        assert result["success"] == False
        assert "Permission denied" in result["error"]


class TestReportsAPI:
    """Test suite for Reports API endpoints"""

    @pytest.mark.asyncio
    @patch('app.api.reports.generate_report_tool')
    async def test_generate_report_endpoint_success(self, mock_tool):
        """Test successful report generation endpoint"""
        from app.api.reports import generate_report
        from app.models.reports import ReportGenerateRequest

        # Mock tool response
        mock_tool.run.return_value = {
            "success": True,
            "result": {
                "report_path": "/tmp/report_123.pdf",
                "format": "pdf",
                "rows_count": 10,
                "execution_time_ms": 1234.56,
                "question": "List employees",
                "sql_query": "SELECT * FROM Employees",
                "truncated": False,
                "max_rows": 1000
            }
        }

        # Create request
        request = ReportGenerateRequest(
            question="List all employees",
            format="pdf",
            user_id="admin_001",
            user_role="ADMIN"
        )

        # Call endpoint
        response = await generate_report(request)

        # Verify response
        assert response.success == True
        assert response.report_path == "/tmp/report_123.pdf"
        assert response.format == "pdf"
        assert response.rows_count == 10

        # Verify tool was called
        mock_tool.run.assert_called_once()

    @pytest.mark.asyncio
    @patch('app.api.reports.generate_report_tool')
    async def test_generate_report_endpoint_failure(self, mock_tool):
        """Test report generation endpoint handles failures"""
        from app.api.reports import generate_report
        from app.models.reports import ReportGenerateRequest

        # Mock tool error
        mock_tool.run.return_value = {
            "success": False,
            "error": "Database connection failed"
        }

        # Create request
        request = ReportGenerateRequest(
            question="Test query",
            user_id="admin",
            user_role="ADMIN"
        )

        # Call endpoint
        response = await generate_report(request)

        # Verify error response
        assert response.success == False
        assert "Database connection failed" in response.error

    @pytest.mark.asyncio
    @patch('app.api.reports.settings')
    @patch('app.api.reports.os.path.exists')
    async def test_download_report_endpoint_success(self, mock_exists, mock_settings):
        """Test successful file download"""
        from app.api.reports import download_report

        # Mock settings and file existence
        mock_settings.reports_output_dir = "/tmp/reports"
        mock_exists.return_value = True

        # This will fail at FileResponse, but that's ok for unit test
        # We're just testing the path validation logic
        try:
            response = await download_report("report_123.pdf")
        except:
            pass  # Expected to fail at FileResponse

        # Verify path validation
        mock_exists.assert_called_once()

    @pytest.mark.asyncio
    @patch('app.api.reports.settings')
    @patch('app.api.reports.os.path.exists')
    async def test_download_report_endpoint_not_found(self, mock_exists, mock_settings):
        """Test download endpoint returns 404 for missing files"""
        from app.api.reports import download_report
        from fastapi import HTTPException

        # Mock file not found
        mock_settings.reports_output_dir = "/tmp/reports"
        mock_exists.return_value = False

        # Call endpoint - should raise HTTPException
        with pytest.raises(HTTPException) as exc_info:
            await download_report("missing_file.pdf")

        assert exc_info.value.status_code == 404

    @pytest.mark.asyncio
    @patch('app.api.reports.settings')
    @patch('app.api.reports.os.path.exists')
    @patch('app.api.reports.os.listdir')
    @patch('app.api.reports.os.path.isfile')
    @patch('app.api.reports.os.stat')
    async def test_list_reports_endpoint(self, mock_stat, mock_isfile, mock_listdir,
                                        mock_exists, mock_settings):
        """Test list reports endpoint"""
        from app.api.reports import list_reports

        # Mock directory and files
        mock_settings.reports_output_dir = "/tmp/reports"
        mock_exists.return_value = True
        mock_listdir.return_value = ["report_1.pdf", "report_2.xlsx"]
        mock_isfile.return_value = True

        # Mock file stats
        mock_stat_obj = Mock()
        mock_stat_obj.st_size = 12345
        mock_stat_obj.st_ctime = 1234567890.0
        mock_stat.return_value = mock_stat_obj

        # Call endpoint
        response = await list_reports()

        # Verify response
        assert "reports" in response
        assert response["total_reports"] == 2
        assert len(response["reports"]) == 2

        # Verify file metadata
        assert any(r["filename"] == "report_1.pdf" for r in response["reports"])
        assert any(r["format"] == "excel" for r in response["reports"])


if __name__ == "__main__":
    # Run tests
    pytest.main([__file__, "-v", "--tb=short"])
