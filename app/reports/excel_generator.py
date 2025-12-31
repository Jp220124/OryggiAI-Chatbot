"""
Excel Report Generator
Uses openpyxl for Excel file generation with formatting
"""

import os
from typing import Dict, Any, Optional, List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger

from app.config import settings
from .report_formatter import ReportFormatter
from .interfaces import ReportGenerator


class ExcelReportGenerator(ReportGenerator):
    """
    Generate Excel reports from query results using openpyxl
    """

    def __init__(self):
        """Initialize Excel generator"""
        # Ensure output directory exists
        os.makedirs(settings.reports_output_dir, exist_ok=True)

        logger.debug("ExcelReportGenerator initialized")

    @property
    def format_name(self) -> str:
        """Return format name for factory registration"""
        return 'excel'

    @property
    def file_extension(self) -> str:
        """Return file extension"""
        return '.xlsx'

    @property
    def mime_type(self) -> str:
        """Return MIME type for HTTP headers"""
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def generate_report(
        self,
        formatted_data: Dict[str, Any],
        filename: Optional[str] = None,
        include_summary: bool = True,
        auto_filter: bool = True,
        freeze_panes: bool = True
    ) -> str:
        """
        Generate Excel report from formatted data

        Args:
            formatted_data: Output from ReportFormatter.format_query_results()
            filename: Output filename (auto-generated if None)
            include_summary: Include summary statistics sheet
            auto_filter: Enable auto-filter on columns
            freeze_panes: Freeze header row

        Returns:
            Path to generated Excel file
        """
        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"report_{timestamp}.xlsx"

            # Ensure .xlsx extension
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'

            output_path = os.path.join(settings.reports_output_dir, filename)

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Create data sheet
            self._create_data_sheet(
                wb=wb,
                formatted_data=formatted_data,
                auto_filter=auto_filter,
                freeze_panes=freeze_panes
            )

            # Create summary sheet if requested
            if include_summary and formatted_data["rows"]:
                df = ReportFormatter.to_dataframe(formatted_data)
                summary_stats = ReportFormatter.create_summary_stats(df)
                self._create_summary_sheet(wb, summary_stats, formatted_data)

            # Create metadata sheet
            if formatted_data.get("metadata"):
                self._create_metadata_sheet(wb, formatted_data["metadata"])

            # Save workbook
            wb.save(output_path)

            logger.info(f"Excel report generated: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Failed to generate Excel report: {str(e)}", exc_info=True)
            raise

    def _create_data_sheet(
        self,
        wb: Workbook,
        formatted_data: Dict[str, Any],
        auto_filter: bool,
        freeze_panes: bool
    ) -> None:
        """
        Create main data sheet

        Args:
            wb: Workbook
            formatted_data: Formatted report data
            auto_filter: Enable auto-filter
            freeze_panes: Freeze header row
        """
        ws = wb.create_sheet("Data")

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        columns = formatted_data.get("columns", [])
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        rows = formatted_data.get("rows", [])
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="top")

        # Auto-size columns
        self._auto_size_columns(ws)

        # Apply auto-filter
        if auto_filter and columns:
            last_col = get_column_letter(len(columns))
            ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

        # Freeze header row
        if freeze_panes:
            ws.freeze_panes = "A2"

        # Add borders
        self._add_borders(ws, len(rows) + 1, len(columns))

    def _create_summary_sheet(
        self,
        wb: Workbook,
        summary_stats: Dict[str, Any],
        formatted_data: Dict[str, Any]
    ) -> None:
        """
        Create summary statistics sheet

        Args:
            wb: Workbook
            summary_stats: Summary statistics
            formatted_data: Formatted report data
        """
        ws = wb.create_sheet("Summary")

        # Title
        ws['A1'] = "Report Summary"
        ws['A1'].font = Font(bold=True, size=14)

        # Basic stats
        row = 3
        ws[f'A{row}'] = "Total Rows:"
        ws[f'B{row}'] = summary_stats.get("total_rows", 0)

        row += 1
        ws[f'A{row}'] = "Total Columns:"
        ws[f'B{row}'] = summary_stats.get("total_columns", 0)

        row += 1
        ws[f'A{row}'] = "Generated At:"
        ws[f'B{row}'] = formatted_data.get("generated_at", "")

        # Numeric summaries
        if summary_stats.get("numeric_summary"):
            row += 2
            ws[f'A{row}'] = "Numeric Column Statistics"
            ws[f'A{row}'].font = Font(bold=True)

            row += 1
            headers = ["Column", "Min", "Max", "Mean", "Median", "Sum"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            for col_name, stats in summary_stats["numeric_summary"].items():
                row += 1
                ws[f'A{row}'] = col_name
                ws[f'B{row}'] = round(stats.get("min", 0), 2)
                ws[f'C{row}'] = round(stats.get("max", 0), 2)
                ws[f'D{row}'] = round(stats.get("mean", 0), 2)
                ws[f'E{row}'] = round(stats.get("median", 0), 2)
                ws[f'F{row}'] = round(stats.get("sum", 0), 2)

        # Auto-size columns
        self._auto_size_columns(ws)

    def _create_metadata_sheet(
        self,
        wb: Workbook,
        metadata: Dict[str, Any]
    ) -> None:
        """
        Create metadata sheet

        Args:
            wb: Workbook
            metadata: Report metadata
        """
        ws = wb.create_sheet("Metadata")

        # Title
        ws['A1'] = "Report Metadata"
        ws['A1'].font = Font(bold=True, size=14)

        # Metadata fields
        row = 3
        for key, value in metadata.items():
            ws[f'A{row}'] = key.replace("_", " ").title()
            ws[f'A{row}'].font = Font(bold=True)

            # Handle SQL query (wrap text)
            if key == "sql_query":
                cell = ws[f'B{row}']
                cell.value = value
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[row].height = 100
            else:
                ws[f'B{row}'] = str(value)

            row += 1

        # Auto-size columns
        self._auto_size_columns(ws)

    def _auto_size_columns(self, ws) -> None:
        """
        Auto-size all columns based on content

        Args:
            ws: Worksheet
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_borders(self, ws, max_row: int, max_col: int) -> None:
        """
        Add borders to all cells

        Args:
            ws: Worksheet
            max_row: Maximum row number
            max_col: Maximum column number
        """
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).border = thin_border

    async def generate_table_report(
        self,
        query_results: List[Dict[str, Any]],
        title: str,
        user_id: str,
        user_role: str,
        question: str,
        sql_query: str,
        filename: Optional[str] = None,
        max_rows: Optional[int] = None
    ) -> str:
        """
        Generate Excel report from raw query results (async for interface compliance)

        Args:
            query_results: Raw query results
            title: Report title
            user_id: User who generated report
            user_role: User's role
            question: Original question
            sql_query: SQL query executed
            filename: Output filename
            max_rows: Maximum rows to include

        Returns:
            Path to generated Excel file
        """
        # Format data
        formatted_data = ReportFormatter.format_query_results(
            query_results=query_results,
            title=title,
            max_rows=max_rows or settings.reports_max_rows
        )

        # Add metadata
        formatted_data = ReportFormatter.add_metadata(
            formatted_data=formatted_data,
            user_id=user_id,
            user_role=user_role,
            question=question,
            sql_query=sql_query
        )

        # Generate report (sync operation, but wrapped in async method)
        return self.generate_report(
            formatted_data=formatted_data,
            filename=filename,
            include_summary=True,
            auto_filter=True,
            freeze_panes=True
        )


# Global instance
excel_generator = ExcelReportGenerator()
